import json
import re
import requests
import concurrent.futures
from bs4 import BeautifulSoup
from datetime import datetime
import time
import os
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import tkinter as tk
from tkinter import scrolledtext
import threading
import sys

# ==============================================================================
# GESTION DES CHEMINS PYINSTALLER
# ==============================================================================
def resource_path(relative_path):
    """ Obtient le chemin absolu vers la ressource, compatible PyInstaller """
    try:
        # PyInstaller crée un dossier temporaire et stocke le chemin dans _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ==============================================================================
# CONFIGURATION DES PÉRIODES DE RÉALISATION
# ==============================================================================
# Format : (Date Début, Date Fin)
PERIODES_MINIMA = {
    "ce":  (datetime(2026, 5, 1), datetime(2026, 7, 26)),
    "u20": (datetime(2026, 4, 1), datetime(2026, 7, 19)),
    "u18": (datetime(2026, 2, 1), datetime(2026, 7, 5))
}

# ==========================================
# CONFIGURATION & CRITÈRES FFA 2026
# ==========================================

# Dictionnaire complété avec les minima issus des fichiers (Minima A, Minima B)
# Lorsqu'il n'y a qu'un minima (comme les U18), la valeur est (Minima, None)
MINIMA_FFA = {
    "ce": {
        "m": {
            "100m": (10.05, 10.15), "200m": (20.20, 20.45), "400m": (45.00, 45.25), "800m": (103.00, 104.80),
            "1500m": (210.50, 213.50), "5000m": (776.00, None), "10000m": (1650.00, None), "Marathon": (7590.00, None),
            "110mH": (13.20, 13.40), "400mH": (48.20, 49.00), "3000m Steeple": (491.00, 500.00), "Hauteur": (2.27, 2.24),
            "Perche": (5.82, 5.70), "Longueur": (8.20, 8.05), "Triple": (17.20, 16.80), "Poids": (21.30, 20.80),
            "Disque": (68.50, 65.00), "Marteau": (80.00, 77.00), "Javelot": (85.00, 83.00), "Decathlon": (8400, 8200),
            "20km Marche": (4800.00, None),
            "Semi-marathon Marche": (5070.00, None),
            "35km Marche": (8760.00, None),
            "Marathon Marche": (10920.00, None)
        },
        "f": {
            "100m": (11.00, 11.18), "200m": (22.50, 22.85), "400m": (50.20, 51.20), "800m": (118.00, 119.80),
            "1500m": (238.00, 243.50), "5000m": (880.00, None), "10000m": (1850.00, None), "Marathon": (8610.00, None),
            "100mH": (12.60, 12.88), "400mH": (54.30, 55.30), "3000m Steeple": (552.00, 568.00), "Hauteur": (1.94, 1.92),
            "Perche": (4.70, 4.60), "Longueur": (6.77, 6.70), "Triple": (14.45, 14.20), "Poids": (19.00, 18.20),
            "Disque": (65.50, 61.00), "Marteau": (73.50, 71.50), "Javelot": (65.50, 60.80), "Heptathlon": (6570, 6250),
            "20km Marche": (5340.00, None),
            "Semi-marathon Marche": (5640.00, None),
            "35km Marche": (9780.00, None),
            "Marathon Marche": (12450.00, None)
        }
    },
    "u20": {
        "m": {
            "100m": (10.20, 10.35), "200m": (20.50, 20.95), "400m": (45.95, 46.80), "800m": (106.80, 107.50),
            "1500m": (217.50, 221.80), "3000m": (467.00, 484.00), "5000m": (810.00, 844.00), "110mH": (13.35, 13.55),
            "400mH": (50.30, 51.20), "3000m Steeple": (520.00, 532.00), "Hauteur": (2.20, 2.15), "Perche": (5.45, 5.25),
            "Longueur": (7.85, 7.65), "Triple": (16.20, 15.80), "Poids": (20.20, 19.20), "Disque": (61.00, 58.50),
            "Marteau": (76.00, 71.00), "Javelot": (75.50, 71.50), "Decathlon": (7550, 7350), "5000m Marche": (1160.00, 1210.00)
        },
        "f": {
            "100m": (11.30, 11.50), "200m": (23.20, 23.65), "400m": (52.65, 53.65), "800m": (122.40, 125.20),
            "1500m": (250.50, 259.00), "3000m": (545.50, 560.00), "5000m": (948.00, 970.00), "100mH": (13.10, 13.55),
            "400mH": (58.00, 58.90), "3000m Steeple": (590.00, 622.00), "Hauteur": (1.88, 1.82), "Perche": (4.35, 4.15),
            "Longueur": (6.45, 6.30), "Triple": (13.40, 13.20), "Poids": (16.25, 15.30), "Disque": (53.50, 51.50),
            "Marteau": (62.60, 61.00), "Javelot": (56.00, 52.00), "Heptathlon": (5750, 5400), "5000m Marche": (1310.00, 1360.00)
        }
    },
    "u18": {
        "m": {
            "100m": (10.65, None), "200m": (21.55, None), "400m": (48.30, None), "800m": (111.50, None),
            "1500m": (231.00, None), "3000m": (501.00, None), "110mH": (13.80, None), "400mH": (53.50, None),
            "2000m Steeple": (352.00, None), "Hauteur": (2.05, None), "Perche": (4.85, None), "Longueur": (7.30, None),
            "Triple": (14.80, None), "Poids": (18.40, None), "Disque": (56.50, None), "Marteau": (69.50, None),
            "Javelot": (68.50, None), "Decathlon": (7000, None), "5000m Marche": (1305.00, None)
        },
        "f": {
            "100m": (11.75, None), "200m": (24.10, None), "400m": (54.80, None), "800m": (126.50, None),
            "1500m": (264.00, None), "3000m": (577.50, None), "100mH": (13.60, None), "400mH": (60.40, None),
            "2000m Steeple": (410.00, None), "Hauteur": (1.79, None), "Perche": (3.90, None), "Longueur": (6.05, None),
            "Triple": (12.70, None), "Poids": (15.85, None), "Disque": (45.60, None), "Marteau": (63.80, None),
            "Javelot": (50.00, None), "Heptathlon": (5400, None), "5000m Marche": (1425.00, None)
        }
    }
}

# Mapping mis à jour pour la recherche par slugs sur worldathletics.org
MAP_WA_SLUGS = {
    "100m": "sprints/100-metres",
    "200m": "sprints/200-metres",
    "400m": "sprints/400-metres",
    "800m": "middlelong/800-metres",
    "1500m": "middlelong/1500-metres",
    "3000m": "middlelong/3000-metres",
    "5000m": "middlelong/5000-metres",
    "10000m": "middlelong/10000-metres",
    "Marathon": "road-running/marathon",
    "100mH": "hurdles/100-metres-hurdles",
    "110mH": "hurdles/110-metres-hurdles",
    "400mH": "hurdles/400-metres-hurdles",
    "2000m Steeple": "middlelong/2000-metres-steeplechase",
    "3000m Steeple": "middlelong/3000-metres-steeplechase",
    "Hauteur": "jumps/high-jump",
    "Perche": "jumps/pole-vault",
    "Longueur": "jumps/long-jump",
    "Triple": "jumps/triple-jump",
    "Poids": "throws/shot-put",
    "Disque": "throws/discus-throw",
    "Marteau": "throws/hammer-throw",
    "Javelot": "throws/javelin-throw",
    "Heptathlon": "combined-events/heptathlon",
    "Decathlon": "combined-events/decathlon",
    "5000m Marche": "race-walks/5000-metres-race-walk",
    "10000m Marche": "race-walks/10000-metres-race-walk",
    "20km Marche": "race-walks/20-kilometres-race-walk",
    "Semi-marathon Marche": "race-walks/half-marathon-race-walk",
    "35km Marche": "race-walks/35-kilometres-race-walk",
    "Marathon Marche": "race-walks/marathon-race-walk"
}

MAP_WA_SLUGS_OVERRIDES = {
    # ── U20 Hommes ──────────────────────────────────────────────────────────
    ("u20", "m", "110mH"):    "hurdles/110m-hurdles-990cm",
    ("u20", "m", "Poids"):    "throws/shot-put-6kg",
    ("u20", "m", "Disque"):   "throws/discus-throw-1750kg",
    ("u20", "m", "Marteau"):  "throws/hammer-throw-6kg",
    ("u20", "m", "Decathlon"):"combined-events/decathlon-u20",

    # ── U18 Hommes ──────────────────────────────────────────────────────────
    ("u18", "m", "110mH"):    "hurdles/110m-hurdles-914cm",
    ("u18", "m", "400mH"):    "hurdles/400m-hurdles-840cm",
    ("u18", "m", "Poids"):    "throws/shot-put-5kg",
    ("u18", "m", "Disque"):   "throws/discus-throw-1500kg",
    ("u18", "m", "Marteau"):  "throws/hammer-throw-5kg",
    ("u18", "m", "Javelot"):  "throws/javelin-throw-700g",
    ("u18", "m", "Decathlon"):"combined-events/decathlon-boys",

    # ── U18 Femmes ──────────────────────────────────────────────────────────
    ("u18", "f", "100mH"):    "hurdles/100m-hurdles-762cm",
    ("u18", "f", "Poids"):    "throws/shot-put-3kg",
    ("u18", "f", "Marteau"):  "throws/hammer-throw-3kg",
    ("u18", "f", "Javelot"):  "throws/javelin-throw-500g",
    ("u18", "f", "Heptathlon"):"combined-events/heptathlon-girls",
}

def get_wa_slug(champ, gender, event):
    """Retourne le slug WA approprié en tenant compte des variantes d'engins par catégorie."""
    override = MAP_WA_SLUGS_OVERRIDES.get((champ, gender, event))
    if override:
        return override
    return MAP_WA_SLUGS.get(event)

def parse_wa_date(date_str):
    """Convertit une date World Athletics (ex: 15 MAY 2026) en objet datetime."""
    try:
        clean_date = date_str.replace(",", "").strip()
        return datetime.strptime(clean_date, "%d %b %Y")
    except Exception:
        return None

MOIS_FR = {
    "JAN": "jan", "FEB": "fév", "MAR": "mar", "APR": "avr",
    "MAY": "mai", "JUN": "juin", "JUL": "juil", "AUG": "août",
    "SEP": "sep", "OCT": "oct", "NOV": "nov", "DEC": "déc"
}

def parse_date_universal(date_str):
    """Convertit une date WA (15 MAY 2026) ou FFA (15/05/26) en datetime."""
    if not date_str:
        return None
    try:
        clean_date = str(date_str).replace(",", "").strip()
        
        # Format FFA (DD/MM/YY ou DD/MM/YYYY)
        m_ffa = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})$', clean_date)
        if m_ffa:
            d_num, m_num, y_num = int(m_ffa.group(1)), int(m_ffa.group(2)), int(m_ffa.group(3))
            y_num = 2000 + y_num if y_num < 100 else y_num
            return datetime(y_num, m_num, d_num)
            
        # Format WA (15 MAY 2026)
        return datetime.strptime(clean_date, "%d %b %Y")
    except Exception:
        return None
        
def translate_date_fr(date_str):
    """Traduit une date WA (ex: 17 MAY 2026) en français (ex: 17 mai 2026)."""
    if not date_str:
        return date_str
    parts = date_str.strip().split()
    if len(parts) == 3:
        mois_en = parts[1].upper()
        if mois_en in MOIS_FR:
            return f"{parts[0]} {MOIS_FR[mois_en]} {parts[2]}"
    return date_str

def is_perf_in_period(date_str, champ):
    """Vérifie si la date de performance est comprise dans la période définie."""
    if champ not in PERIODES_MINIMA:
        return True
    
    perf_date = parse_date_universal(date_str)
    if not perf_date:
        return False
        
    start, end = PERIODES_MINIMA[champ]
    return start <= perf_date <= end

def time_to_seconds(time_str):
    """Parseur robuste pour comprendre les formats bruts."""
    if isinstance(time_str, (int, float)): return float(time_str)
    
    time_str = str(time_str).strip().lower().replace(',', '.')
    if not time_str or time_str == "-": return None
    
    # Séparer les +, -, et autres annotations (exclut les (w))
    time_str = re.split(r'\s*\+|\s*\-|\s*\(', time_str)[0].strip()
    
    # SUPPRESSION CRITIQUE : enlever le 'i' (indoor), 'a' (altitude), et '*' des perfs Tilastopaja
    time_str = re.sub(r'[a-z\*\+]+$', '', time_str).strip()
    
    try:
        if 'm' in time_str and 'h' not in time_str and "'" not in time_str:
            parts = time_str.split('m')
            m = float(parts[0]) if parts[0] else 0
            cm = float(parts[1]) if len(parts) > 1 and parts[1] else 0
            return float(f"{int(m)}.{int(cm):02d}")
            
        if re.match(r"^\d+(\.\d+)?$", time_str):
            return float(time_str)

        if ':' in time_str:
            parts = time_str.split(':')
            if len(parts) == 3:
                return float(parts[0])*3600 + float(parts[1])*60 + float(parts[2])
            elif len(parts) == 2:
                return float(parts[0])*60 + float(parts[1])
            
        time_str = time_str.replace("''", '"').replace("’", "'").replace("´", "'")
        h, m, s, c = 0, 0, 0, 0
        
        if 'h' in time_str:
            parts = time_str.split('h')
            h = float(parts[0])
            time_str = parts[1]
            
        if "'" in time_str:
            parts = time_str.split("'")
            m = float(parts[0]) if parts[0] else 0
            time_str = parts[1]
            
        if '"' in time_str:
            parts = time_str.split('"')
            s = float(parts[0]) if parts[0] else 0
            c = float(parts[1]) if len(parts) > 1 and parts[1] else 0
        elif time_str:
            s = float(time_str)
                
        return h * 3600 + m * 60 + s + c / 100.0

    except Exception:
        return None

def format_seconds_for_display(seconds, event_name):
    """Formate les performances selon l'épreuve."""
    if seconds is None or seconds == "-": return "-"
    
    event_lower = event_name.lower()
    
    if "heptathlon" in event_lower or "decathlon" in event_lower:
        return str(int(seconds))
        
    field_events = ["hauteur", "perche", "longueur", "triple", "poids", "disque", "marteau", "javelot"]
    if any(f in event_lower for f in field_events):
        m_val = int(seconds)
        cm_val = int(round((seconds - m_val) * 100))
        return f"{m_val}m{cm_val:02d}"
        
    short_events = ["100m", "200m", "400m", "110mh", "100mh", "110m haies", "100m haies", "400mh", "400m haies", "100m/110m haies"]
    is_short = any((event_lower == se or event_lower.startswith(se + " ")) for se in short_events)
            
    if is_short:
        s_val = int(seconds)
        c_val = int(round((seconds - s_val) * 100))
        return f"{s_val}\"{c_val:02d}"
        
    h_val = int(seconds // 3600)
    m_val = int((seconds % 3600) // 60)
    s_val = int(seconds % 60)
    c_val = int(round((seconds - (h_val*3600 + m_val*60 + s_val)) * 100))
    
    if h_val > 0:
        return f"{h_val}h{m_val:02d}'{s_val:02d}\"{c_val:02d}"
    else:
        return f"{m_val}'{s_val:02d}\"{c_val:02d}"
        
# ==============================================================================
# SCRAPING FFA (Remplace Tilastopaja)
# ==============================================================================

# Dictionnaire des codes épreuves FFA. 
MAP_FFA_EPREUVES = {
    "100m": "110", "200m": "120", "400m": "130",
    "800m": "140", "1500m": "150", "3000m": "160",
    "5000m": "170", "10000m": "180", "Marathon": "205",
    "100mH": "210", "110mH": "210", "400mH": "220",
    "2000m Steeple": "260", "3000m Steeple": "250",
    "Hauteur": "310", "Perche": "320", "Longueur": "330", "Triple": "340",
    "Poids": "410", "Disque": "420", "Marteau": "430", "Javelot": "440",
    "Heptathlon": "570", "Decathlon": "580",
    "5000m Marche": "610", "10000m Marche": "620", "20km Marche": "630",
    "35km Marche": "640", "Semi-marathon Marche": "635", "Marathon Marche": "645"
}

def fetch_ffa_event(champ, gender, event):
    """Scrape les bilans officiels de la FFA pour une épreuve précise."""
    epreuve_id = MAP_FFA_EPREUVES.get(event)
    if not epreuve_id:
        return []

    # Le champ vide ("") pour 'ce' (Seniors/Europe) permet de prendre toutes les catégories
    cat_map = {"ce": "", "u20": "JU", "u18": "CA"}
    sexe_map = {"m": "M", "f": "F"}
    
    cat = cat_map.get(champ, "")
    sexe = sexe_map.get(gender, "M")

    # frmnationalite=1 garantit que seuls les athlètes français sont pris en compte
    url = f"https://www.athle.fr/bases/liste.aspx?frmpostback=true&frmbase=bilans&frmmode=1&frmespace=0&frmannee=2026&frmepreuve={epreuve_id}&frmsexe={sexe}&frmcategorie={cat}&frmdepartement=&frmligue=&frmnationalite=1&frmvent=VR&frmamaxi="

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml"
    }
    
    athletes = []
    try:
        res = requests.get(url, headers=headers, timeout=15)
        if res.status_code != 200: return []
        
        soup = BeautifulSoup(res.content, 'html.parser')
        
        table = soup.find('table', id='ctnBilans')
        if not table: return []
        
        # On parcourt toutes les lignes du tableau
        for tr in table.find_all('tr'):
            # La nouvelle structure FFA double chaque ligne pour la version mobile ("detail-row"), on l'ignore.
            if 'detail-row' in tr.get('class', []):
                continue
                
            tds = tr.find_all('td')
            # Une ligne de résultat valide possède au moins 9 colonnes (Place, Perf, Nom, Club, Ligue, Dep, Infos, Date, Lieu)
            if len(tds) < 9: continue
            
            # Colonne 2 (index 1) : La perf. 
            # On remplace les doubles apostrophes par des guillemets et on supprime le vent ou la mention (RP)
            # Ex: "10''18 (+2.0) (RP)" devient "10"18"
            perf_raw = tds[1].get_text(strip=True).replace("''", '"')
            perf_text = perf_raw.split('(')[0].strip()
            if not re.search(r'\d', perf_text): continue
            
            # Colonne 3 (index 2) : Nom de l'athlète situé dans le lien
            a_tag = tds[2].find('a')
            if not a_tag: continue
            name_clean = " ".join([w.capitalize() for w in a_tag.get_text(strip=True).split()])
            
            # Colonne 8 (index 7) : Date
            date_text = tds[7].get_text(strip=True)
            
            # Colonne 9 (index 8) : Lieu
            venue_text = tds[8].get_text(strip=True)
            
            athletes.append({
                "name": name_clean,
                "perf": perf_text,
                "date": date_text,
                "place": venue_text,
                "raw_date": date_text
            })
    except Exception as e:
        print(f"❌ Erreur scraping FFA pour {champ.upper()} - {event}: {e}")
        
    return athletes

def merge_and_filter_athletes(wa_list, ffa_list, event_name, limit_to_check, champ):
    """Fusionne intelligemment les deux listes, filtre par minima et période, et ne garde que la meilleure perf."""
    unique_athletes = {}
    
    is_running = event_name in ["100m", "200m", "400m", "800m", "1500m", "3000m", "5000m", "10000m", "100mH", "110mH", "400mH", "2000m Steeple", "3000m Steeple", "5000m Marche", "10000m Marche", "20km Marche", "Semi-marathon Marche", "35km Marche", "Marathon Marche", "Marathon"]
    
    for ath in wa_list + ffa_list:
        name = ath["name"].strip()
        name_lower = name.lower()
        perf_v = time_to_seconds(ath["perf"])
        if perf_v is None: continue
        
        # Filtrage par Minima
        if is_running:
            if perf_v > limit_to_check: continue
        else:
            if perf_v < limit_to_check: continue
            
        # Filtrage par Date (période de réalisation)
        date_str = ath.get("raw_date", ath.get("date"))
        if not is_perf_in_period(date_str, champ):
            continue
            
        # Conservation de la meilleure performance uniquement
        if name_lower not in unique_athletes:
            unique_athletes[name_lower] = ath
        else:
            existing_perf_v = time_to_seconds(unique_athletes[name_lower]["perf"])
            if is_running:
                if perf_v < existing_perf_v: unique_athletes[name_lower] = ath
            else:
                if perf_v > existing_perf_v: unique_athletes[name_lower] = ath
                
    # Trie des résultats finaux (du meilleur au moins bon)
    results_list = list(unique_athletes.values())
    results_list.sort(key=lambda x: time_to_seconds(x["perf"]) or 999999 if is_running else -(time_to_seconds(x["perf"]) or 0))
    
    return results_list

# ==============================================================================
# SCRAPING WORLD ATHLETICS
# ==============================================================================

def fetch_wa_event(champ, gender, event):
    """Scrape le site de World Athletics et filtre strictement par catégorie d'âge, lieu et date."""
    slug = get_wa_slug(champ, gender, event)
    if not slug: return []

    wa_gender = "women" if gender == "f" else "men"
    wa_category = "senior"
    if champ == "u18": wa_category = "u18"
    elif champ == "u20": wa_category = "u20"

    year = "2026"
    url = f"https://worldathletics.org/records/toplists/{slug}/all/{wa_gender}/{wa_category}/{year}?regionType=countries&region=fra&timing=electronic&windReading=regular&page=1&bestResultsOnly=true"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "fr,fr-FR;q=0.8,en-US;q=0.5,en;q=0.3"
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=15)
        if response.status_code != 200: return []
        
        soup = BeautifulSoup(response.content, 'html.parser')
        athletes = []
        
        limit_tuple = MINIMA_FFA[champ][gender][event]
        limit_a, limit_b = limit_tuple[0], limit_tuple[1]
        limit_to_check = limit_b if limit_b is not None else limit_a
        
        is_running = event in ["100m", "200m", "400m", "800m", "1500m", "3000m", "5000m", "10000m", "100mH", "110mH", "400mH", "2000m Steeple", "3000m Steeple", "5000m Marche", "10000m Marche", "20km Marche", "Semi-marathon Marche", "35km Marche", "Marathon Marche", "Marathon"]
        
        table = soup.find('table', {'class': 'records-table'})
        if not table: return []

        # Cartographie intelligente des colonnes pour trouver DOB, Lieu et Date
        col_map = {}
        thead = table.find('thead')
        if thead:
            for i, th in enumerate(thead.find_all(['th', 'td'])):
                col_map[th.text.strip().lower()] = i

        tbody = table.find('tbody')
        rows = tbody.find_all('tr') if tbody else table.find_all('tr')[1:]

        for row in rows:
            cols = row.find_all('td')
            if len(cols) < 4: continue
            
            # Identifier la colonne de la Performance
            mark_col = cols[col_map['mark']] if 'mark' in col_map and col_map['mark'] < len(cols) else cols[1]
            
            # Identifier la colonne de l'athlète
            competitor_col = None
            if 'competitor' in col_map and col_map['competitor'] < len(cols):
                competitor_col = cols[col_map['competitor']]
            else:
                for c in cols:
                    if c.find('a', href=re.compile(r"/athletes/")):
                        competitor_col = c
                        break
            if not competitor_col: continue
            
            # Récupérer Lieu (Venue) et Date
            venue_text = cols[col_map['venue']].text.strip() if 'venue' in col_map and col_map['venue'] < len(cols) else cols[-2].text.strip()
            date_text = cols[col_map['date']].text.strip() if 'date' in col_map and col_map['date'] < len(cols) else cols[-1].text.strip()

            # FILTRAGE PAR PÉRIODE DE RÉALISATION
            if not is_perf_in_period(date_text, champ):
                continue

            # Récupérer l'année de naissance pour un filtrage strict (DOB)
            dob_text = ""
            if 'dob' in col_map and col_map['dob'] < len(cols):
                dob_text = cols[col_map['dob']].text.strip()
            else:
                comp_idx = cols.index(competitor_col) if competitor_col in cols else -1
                if comp_idx != -1 and comp_idx + 1 < len(cols):
                    dob_text = cols[comp_idx + 1].text.strip()

            # Filtrage EXCLUSIF des catégories (U18 = 2009-2010 | U20 = 2007-2008)
            if champ in ['u18', 'u20']:
                dob_match = re.search(r'\b(19|20)\d{2}\b', dob_text)
                if dob_match:
                    dob_year = int(dob_match.group(0))
                    if champ == 'u18' and dob_year not in [2009, 2010]:
                        continue
                    if champ == 'u20' and dob_year not in [2007, 2008]:
                        continue

            name_text = competitor_col.text.strip()
            perf_text = mark_col.text.strip().replace('A', '').strip()

            perf_val = time_to_seconds(perf_text)
            if perf_val is None: continue
                
            is_qualified = False
            if is_running:
                if perf_val <= limit_to_check: is_qualified = True
            else:
                if perf_val >= limit_to_check: is_qualified = True
                    
            if is_qualified:
                name_clean = " ".join([w.capitalize() for w in name_text.split()])
                athletes.append({
                    "name": name_clean, 
                    "perf": perf_text, 
                    "date": translate_date_fr(date_text), 
                    "place": venue_text,
                    "raw_date": date_text
                })
        
        # Déduplication et conservation de la meilleure perf
        unique_athletes = {}
        for ath in athletes:
            name = ath["name"]
            perf_v = time_to_seconds(ath["perf"])
            if name not in unique_athletes:
                unique_athletes[name] = ath
            else:
                existing_perf_v = time_to_seconds(unique_athletes[name]["perf"])
                if is_running:
                    if perf_v < existing_perf_v: unique_athletes[name] = ath
                else:
                    if perf_v > existing_perf_v: unique_athletes[name] = ath
                        
        return list(unique_athletes.values())

    except Exception as e:
        print(f"❌ Erreur lors de la récupération de {champ}-{gender}-{event}: {e}")
        return []

# ==========================================
# GESTION DES CORRESPONDANCES & DES STYLES EXCEL
# ==========================================
def get_event_keys(event_name_raw):
    """Associe l'intitulé d'une ligne Excel aux clés correspondantes du scraper."""
    name = str(event_name_raw).strip().lower()
    name = re.sub(r'\s+', ' ', name)
    name = name.replace('ê', 'e').replace('é', 'e').replace('è', 'e')
    
    femmes_key, hommes_key = None, None
    if "100m/110m" in name: femmes_key, hommes_key = "100mH", "110mH"
    elif "100m haies" in name or name == "100mh" or name.startswith("100m haies"): femmes_key = "100mH"
    elif "110m haies" in name or name == "110mh" or name.startswith("110m haies"): hommes_key = "110mH"
    elif "400m haies" in name or name == "400mh": femmes_key, hommes_key = "400mH", "400mH"
    elif "2000m steeple" in name or "2000s" in name: femmes_key, hommes_key = "2000m Steeple", "2000m Steeple"
    elif "3000m steeple" in name or "3000s" in name: femmes_key, hommes_key = "3000m Steeple", "3000m Steeple"
    elif "semi-marathon marche" in name or "semi marathon marche" in name: femmes_key, hommes_key = "Semi-marathon Marche", "Semi-marathon Marche"
    elif "35km marche" in name or "35 km marche" in name: femmes_key, hommes_key = "35km Marche", "35km Marche"
    elif "marathon marche" in name: femmes_key, hommes_key = "Marathon Marche", "Marathon Marche"
    elif "10000m" in name or "10 000m" in name:
        if "marche" in name: femmes_key, hommes_key = "10000m Marche", "10000m Marche"
        else: femmes_key, hommes_key = "10000m", "10000m"
    elif "5000m" in name or "5 000m" in name:
        if "marche" in name: femmes_key, hommes_key = "5000m Marche", "5000m Marche"
        else: femmes_key, hommes_key = "5000m", "5000m"
    elif "20km marche" in name or "20 km marche" in name: femmes_key, hommes_key = "20km Marche", "20km Marche"
    elif "marathon" in name: femmes_key, hommes_key = "Marathon", "Marathon"
    elif "100m" in name: femmes_key, hommes_key = "100m", "100m"
    elif "200m" in name: femmes_key, hommes_key = "200m", "200m"
    elif "400m" in name: femmes_key, hommes_key = "400m", "400m"
    elif "800m" in name: femmes_key, hommes_key = "800m", "800m"
    elif "1500m" in name: femmes_key, hommes_key = "1500m", "1500m"
    elif "3000m" in name: femmes_key, hommes_key = "3000m", "3000m"
    elif "hauteur" in name: femmes_key, hommes_key = "Hauteur", "Hauteur"
    elif "perche" in name: femmes_key, hommes_key = "Perche", "Perche"
    elif "longueur" in name: femmes_key, hommes_key = "Longueur", "Longueur"
    elif "triple" in name: femmes_key, hommes_key = "Triple", "Triple"
    elif "poids" in name: femmes_key, hommes_key = "Poids", "Poids"
    elif "disque" in name: femmes_key, hommes_key = "Disque", "Disque"
    elif "marteau" in name: femmes_key, hommes_key = "Marteau", "Marteau"
    elif "javelot" in name: femmes_key, hommes_key = "Javelot", "Javelot"
    elif "heptathlon" in name: femmes_key = "Heptathlon"
    elif "decathlon" in name: hommes_key = "Decathlon"
    elif "combin" in name: femmes_key, hommes_key = "Heptathlon", "Decathlon"
    return femmes_key, hommes_key

def get_champ_key_from_sheet_name(sheet_name):
    s = sheet_name.lower()
    if "u18" in s or "rieti" in s: return "u18"
    elif "u20" in s or "eugene" in s: return "u20"
    elif "europe" in s or "birmingham" in s or "ce" in s: return "ce"
    return None

def create_standalone_template(wb_path):
    print(f"🪄 Génération automatique du fichier modèle avec la gestion des colonnes Minima A/B...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active) 
    
    champ_titles = {
        "ce": "CHAMPIONNATS D'EUROPE - Birmingham - 10 au 16 août 2026",
        "u20": "CHAMPIONNATS DU MONDE U20 - Eugene (USA) - 05 au 09 août 2026",
        "u18": "CHAMPIONNATS D'EUROPE U18 - Rieti (ITA) - 16 au 19 juillet 2026"
    }
    sheet_names = {"ce": "Europe - Birmingham", "u20": "Monde U20 - Eugene", "u18": "Europe U18 - Rieti"}
    
    ordered_events = [
        "100m", "200m", "400m", "800m", "1500m", "3000m", "5000m", "10000m", "Marathon",
        "100mH", "110mH", "400mH", "2000m Steeple", "3000m Steeple",
        "Hauteur", "Perche", "Longueur", "Triple", "Poids", "Disque", "Marteau", "Javelot",
        "Heptathlon", "Decathlon", "5000m Marche", "10000m Marche", "20km Marche", 
        "Semi-marathon Marche", "35km Marche", "Marathon Marche"
    ]

    for champ_key, sheet_title in sheet_names.items():
        ws = wb.create_sheet(title=sheet_title)
        ws.append([champ_titles[champ_key]])
        
        if champ_key in ["ce", "u20"]:
            ws.append(["Athlètes l'ayant réalisé", "Femmes", "", "ÉPREUVES", "Hommes", "", "Athlètes l'ayant réalisé"])
            ws.append(["", "Minima A", "Minima B", "", "Minima A", "Minima B", ""])
        else:
            ws.append(["", "", "", "", ""])
            ws.append(["Athlète l'ayant réalisé", "Femmes", "Épreuve", "Hommes", "Athlète l'ayant réalisé"])

        for ev in ordered_events:
            f_val_tuple = MINIMA_FFA[champ_key]['f'].get(ev)
            m_val_tuple = MINIMA_FFA[champ_key]['m'].get(ev)
            
            if f_val_tuple is None and m_val_tuple is None:
                continue
            
            ev_display = ev
            if ev == "100mH": ev_display = "100m haies"
            if ev == "110mH": ev_display = "110m haies"
            if ev == "400mH": ev_display = "400m haies"
            if ev == "Triple": ev_display = "Triple Saut"

            if champ_key in ["ce", "u20"]:
                f_a = format_seconds_for_display(f_val_tuple[0] if f_val_tuple else None, ev_display)
                f_b = format_seconds_for_display(f_val_tuple[1] if f_val_tuple and len(f_val_tuple) > 1 else None, ev_display)
                m_a = format_seconds_for_display(m_val_tuple[0] if m_val_tuple else None, ev_display)
                m_b = format_seconds_for_display(m_val_tuple[1] if m_val_tuple and len(m_val_tuple) > 1 else None, ev_display)
                ws.append(["", f_a, f_b, ev_display, m_a, m_b, ""])
            else:
                f_a = format_seconds_for_display(f_val_tuple[0] if f_val_tuple else None, ev_display)
                m_a = format_seconds_for_display(m_val_tuple[0] if m_val_tuple else None, ev_display)
                ws.append(["", f_a, ev_display, m_a, ""])

    wb.save(wb_path)

def set_column_widths(sheet, champ_key):
    if champ_key in ["ce", "u20"]:
        widths = {1: 45, 2: 12, 3: 12, 4: 18, 5: 12, 6: 12, 7: 45} # Colonnes plus larges pour accueillir le lieu et la date
    else: 
        widths = {1: 45, 2: 14, 3: 18, 4: 14, 5: 45}
    for col_idx, width in widths.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = width

def beautify_and_format_sheet(sheet, champ_key):
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.paperSize = 9 
    sheet.page_setup.fitToHeight = 1
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.views.sheetView[0].showGridLines = True
    
    font_name = "Segoe UI"
    title_font = Font(name=font_name, size=14, bold=True, color="1E3A8A")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    sub_header_font = Font(name=font_name, size=9, bold=True, color="475569")
    data_font = Font(name=font_name, size=9, color="0F172A")
    event_font = Font(name=font_name, size=9, bold=True, color="1E3A8A")
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    zebra_even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    zebra_odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    title_bottom_border = Border(bottom=Side(border_style="medium", color="1E3A8A"))
    
    max_col = 7 if champ_key in ["ce", "u20"] else 5
    start_row = 4 if champ_key in ["ce", "u20"] else 3
    
    sheet.row_dimensions[1].height = 40
    is_merged = False
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.start_row == 1 and merged_range.start_column == 1:
            is_merged = True; break
    if not is_merged:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        cell.border = title_bottom_border
        if col == 1:
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
    if champ_key in ["ce", "u20"]:
        sheet.row_dimensions[2].height = 26
        sheet.row_dimensions[3].height = 18
        for col in range(1, max_col + 1):
            cell_r2 = sheet.cell(row=2, column=col)
            cell_r2.fill = header_fill
            cell_r2.font = header_font
            cell_r2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_r2.border = Border(left=Side(border_style="thin", color="38BDF8"), right=Side(border_style="thin", color="38BDF8"), top=thin_border_side, bottom=thin_border_side)
            cell_r3 = sheet.cell(row=3, column=col)
            val = str(cell_r3.value or "").strip()
            if "Minima" in val or val == "A" or val == "B" or "-" in val:
                cell_r3.fill = sub_header_fill
                cell_r3.font = sub_header_font
                cell_r3.alignment = Alignment(horizontal="center", vertical="center")
                cell_r3.border = grid_border
            else:
                cell_r3.fill = header_fill
                cell_r3.border = Border(left=Side(border_style="thin", color="38BDF8"), right=Side(border_style="thin", color="38BDF8"), bottom=thin_border_side)
    else:
        sheet.row_dimensions[2].height = 26
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=Side(border_style="thin", color="38BDF8"), right=Side(border_style="thin", color="38BDF8"), top=thin_border_side, bottom=thin_border_side)

    event_col = 4 if champ_key in ["ce", "u20"] else 3
    femmes_ath_col = 1
    hommes_ath_col = 7 if champ_key in ["ce", "u20"] else 5
        
    for row_idx in range(start_row, sheet.max_row + 1):
        event_val = sheet.cell(row=row_idx, column=event_col).value
        if not event_val: continue
            
        event_str = str(event_val).strip()
        if not event_str or len(event_str) < 2 or event_str.lower().startswith("épreuve") or event_str.lower().startswith("epreuve"):
            sheet.row_dimensions[row_idx].height = 22
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col)
                cell.fill = sub_header_fill
                cell.font = sub_header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = grid_border
            continue
            
        row_fill = zebra_even_fill if row_idx % 2 == 0 else zebra_odd_fill
        femmes_ath_val = str(sheet.cell(row=row_idx, column=femmes_ath_col).value or "")
        hommes_ath_val = str(sheet.cell(row=row_idx, column=hommes_ath_col).value or "")
        
        # Calcul strict des lignes : on compte les sauts de ligne ET on anticipe les retours à la ligne automatiques
        lines_f = sum(max(1, len(line) // 45 + 1) for line in femmes_ath_val.split("\n")) if femmes_ath_val else 1
        lines_m = sum(max(1, len(line) // 45 + 1) for line in hommes_ath_val.split("\n")) if hommes_ath_val else 1
        max_lines = max(1, lines_f, lines_m)
        
        # Ajustement dynamique strict au contenu (13 points par ligne pour une police taille 9) sans marge de confort superflue
        sheet.row_dimensions[row_idx].height = max(18, 13 * max_lines)
        
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = grid_border
            if col_idx == femmes_ath_col or col_idx == hommes_ath_col:
                cell.font = data_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif col_idx == event_col:
                cell.font = event_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    set_column_widths(sheet, champ_key)

# ==========================================
# FONCTION PRINCIPALE ET APPELS
# ==========================================

def run_scraping():
    """Fonction principale de scraping qui tourne en arrière-plan"""
    print("🚀 Début de la veille globale (World Athletics + FFA)...")
    
    final_data = {"ce": {"m": {}, "f": {}}, "u18": {"m": {}, "f": {}}, "u20": {"m": {}, "f": {}}}
    taches_wa = []
    taches_ffa = []
    
    print("\n🌐 Récupération simultanée WA et FFA et croisement des données...")
    
    # Nous lançons les requêtes en parallèle (5 requêtes à la fois pour ne pas surcharger les serveurs)
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        
        # 1. Planification de toutes les requêtes (WA et FFA)
        for champ in MINIMA_FFA.keys():
            for gender in ["m", "f"]:
                for event in MINIMA_FFA[champ][gender].keys():
                    taches_wa.append((champ, gender, event, executor.submit(fetch_wa_event, champ, gender, event)))
                    taches_ffa.append((champ, gender, event, executor.submit(fetch_ffa_event, champ, gender, event)))

        # 2. Récolte des résultats de World Athletics
        results_wa = {}
        for champ, gender, event, future in taches_wa:
            if champ not in results_wa: results_wa[champ] = {"m":{}, "f":{}}
            results_wa[champ][gender][event] = future.result()
            
        # 3. Récolte des résultats de la FFA
        results_ffa = {}
        for champ, gender, event, future in taches_ffa:
            if champ not in results_ffa: results_ffa[champ] = {"m":{}, "f":{}}
            results_ffa[champ][gender][event] = future.result()
            
        # 4. Fusion et filtrage
        for champ in MINIMA_FFA.keys():
            for gender in ["m", "f"]:
                for event in MINIMA_FFA[champ][gender].keys():
                    wa_res = results_wa.get(champ, {}).get(gender, {}).get(event, [])
                    ffa_res = results_ffa.get(champ, {}).get(gender, {}).get(event, [])
                    
                    limit_tuple = MINIMA_FFA[champ][gender][event]
                    limit_to_check = limit_tuple[1] if limit_tuple[1] is not None else limit_tuple[0]
                    
                    # Fusion intelligente WA + FFA
                    resultats_fusionnes = merge_and_filter_athletes(wa_res, ffa_res, event, limit_to_check, champ)
                    
                    if resultats_fusionnes:
                        final_data[champ][gender][event] = resultats_fusionnes
                        print(f"✅ {len(resultats_fusionnes)} qualifié(s) (WA+FFA) pour {champ.upper()} - {gender.upper()} - {event}")

    contenu_js = f"const achievements = {json.dumps(final_data, indent=2, ensure_ascii=False)};"
    with open("achievements.js", "w", encoding="utf-8") as f: f.write(contenu_js)
    
    wb_path = "Minima été 2026.xlsx"
    if not os.path.exists(wb_path):
        create_standalone_template(wb_path)

    print("\n✍️  Mise à jour et formatage graphique du fichier Excel...")
    wb = openpyxl.load_workbook(wb_path)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        champ_key = get_champ_key_from_sheet_name(sheet_name)
        if not champ_key: continue
            
        print(f"📊 Traitement de l'onglet : {sheet_name} ({champ_key.upper()})")
        
        if champ_key in ["ce", "u20"]:
            event_col, femmes_ath_col, hommes_ath_col, start_row = 4, 1, 7, 4
        else:
            event_col, femmes_ath_col, hommes_ath_col, start_row = 3, 1, 5, 3
            
        for row_idx in range(start_row, sheet.max_row + 1):
            event_val = sheet.cell(row=row_idx, column=event_col).value
            if not event_val: continue
                
            event_str = str(event_val).strip()
            if not event_str or len(event_str) < 2 or event_str.lower().startswith("épreuve"): continue
                
            femmes_key, hommes_key = get_event_keys(event_str)
            
            if champ_key in ["ce", "u20"]:
                has_femmes_event = any(v and v != "-" for v in (str(sheet.cell(row=row_idx, column=2).value or "").strip(), str(sheet.cell(row=row_idx, column=3).value or "").strip()))
                has_hommes_event = any(v and v != "-" for v in (str(sheet.cell(row=row_idx, column=5).value or "").strip(), str(sheet.cell(row=row_idx, column=6).value or "").strip()))
            else:
                v_f = str(sheet.cell(row=row_idx, column=2).value or "").strip()
                has_femmes_event = v_f and v_f != "-"
                v_m = str(sheet.cell(row=row_idx, column=4).value or "").strip()
                has_hommes_event = v_m and v_m != "-"
                
            # Insertion athlètes féminines avec ajout du lieu et de la date
            if has_femmes_event and femmes_key:
                ath_list = final_data.get(champ_key, {}).get("f", {}).get(femmes_key, [])
                if ath_list:
                    formatted_lines = []
                    for a in ath_list:
                        perf_sec = time_to_seconds(a['perf'])
                        display_perf = format_seconds_for_display(perf_sec, femmes_key)
                        place_date_info = []
                        if a.get('place') and str(a['place']).strip(): 
                            place_date_info.append(str(a['place']).strip())
                        if a.get('date') and str(a['date']).strip(): 
                            place_date_info.append(f"le {str(a['date']).strip()}")
                        pd_str = " - " + ", ".join(place_date_info) if place_date_info else ""
                        formatted_lines.append(f"• {a['name']} ({display_perf}{pd_str})")
                    sheet.cell(row=row_idx, column=femmes_ath_col, value="\n".join(formatted_lines))
                else: sheet.cell(row=row_idx, column=femmes_ath_col, value="")
            else: sheet.cell(row=row_idx, column=femmes_ath_col, value="")
                
            # Insertion athlètes masculins avec ajout du lieu et de la date
            if has_hommes_event and hommes_key:
                ath_list = final_data.get(champ_key, {}).get("m", {}).get(hommes_key, [])
                if ath_list:
                    formatted_lines = []
                    for a in ath_list:
                        perf_sec = time_to_seconds(a['perf'])
                        display_perf = format_seconds_for_display(perf_sec, hommes_key)
                        place_date_info = []
                        if a.get('place') and str(a['place']).strip(): 
                            place_date_info.append(str(a['place']).strip())
                        if a.get('date') and str(a['date']).strip(): 
                            place_date_info.append(f"le {str(a['date']).strip()}")
                        pd_str = " - " + ", ".join(place_date_info) if place_date_info else ""
                        formatted_lines.append(f"• {a['name']} ({display_perf}{pd_str})")
                    sheet.cell(row=row_idx, column=hommes_ath_col, value="\n".join(formatted_lines))
                else: sheet.cell(row=row_idx, column=hommes_ath_col, value="")
            else: sheet.cell(row=row_idx, column=hommes_ath_col, value="")
        
        beautify_and_format_sheet(sheet, champ_key)
        
    date_str = datetime.now().strftime("%d%m%Y")
    output_filename = f"Minima réalisés au {date_str}.xlsx"
    
    if os.path.exists(wb_path) and output_filename != wb_path:
        wb.save(output_filename)
        os.remove(wb_path)
    else:
        wb.save(output_filename)
        
    print(f"\n🎉 Succès ! Le fichier Excel a été généré et sauvegardé sous le nom : '{output_filename}'")

# ==========================================
# INTERFACE GRAPHIQUE ET THREADING
# ==========================================
class PrintRedirector:
    """Redirige les print() vers le composant Text de l'interface"""
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        # Utiliser .after pour garantir la sécurité des threads dans Tkinter
        self.text_widget.after(0, self._write_thread_safe, string)

    def _write_thread_safe(self, string):
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)
        self.text_widget.configure(state='disabled')

    def flush(self):
        pass

def start_process(btn_start):
    btn_start.config(state=tk.DISABLED, text="Recherche en cours...")
    
    def thread_target():
        try:
            run_scraping()
        except Exception as e:
            print(f"\n❌ Erreur inattendue : {e}")
        finally:
            print("\n✔️ Processus terminé. Vous pouvez fermer la fenêtre.")
            btn_start.after(0, lambda: btn_start.config(state=tk.NORMAL, text="Lancer la recherche"))

    # Lancement dans un thread séparé (daemon=True pour tuer le thread si on ferme la fenêtre)
    threading.Thread(target=thread_target, daemon=True).start()

def launch_gui():
    root = tk.Tk()
    root.title("Veille Minima FFA 2026 - Outil de Scraping")
    root.geometry("850x600")
    root.configure(bg="#F8FAFC")

    # Tentative de chargement de l'icône compilée avec PyInstaller
    icon_path = resource_path("piste.ico")
    if os.path.exists(icon_path):
        try:
            root.iconbitmap(icon_path)
        except:
            pass

    # Titre en haut
    lbl_title = tk.Label(root, text="Veille Minima FFA 2026", font=("Segoe UI", 16, "bold"), bg="#F8FAFC", fg="#1E3A8A")
    lbl_title.pack(pady=15)

    # Zone de logs
    frame_logs = tk.Frame(root, bg="#CBD5E1", bd=1)
    frame_logs.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
    
    log_area = scrolledtext.ScrolledText(frame_logs, wrap=tk.WORD, font=("Consolas", 10), bg="#0F172A", fg="#E2E8F0")
    log_area.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
    log_area.configure(state='disabled')

    # Redirection de la sortie standard
    sys.stdout = PrintRedirector(log_area)
    sys.stderr = PrintRedirector(log_area)

    print("=== Interface prête ===")
    print("Cliquez sur 'Lancer la recherche' pour démarrer l'extraction des minima réalisés (source : bilans World Athletics).")
    print("Le processus peut prendre quelques minutes, veuillez patienter...\n")

    # Bouton de lancement
    frame_btn = tk.Frame(root, bg="#F8FAFC")
    frame_btn.pack(pady=20)
    
    btn_start = tk.Button(frame_btn, text="Lancer la recherche", font=("Segoe UI", 12, "bold"), 
                          bg="#1E3A8A", fg="white", activebackground="#2563EB", activeforeground="white",
                          padx=20, pady=10, cursor="hand2", relief=tk.FLAT)
    btn_start.config(command=lambda: start_process(btn_start))
    btn_start.pack()

    root.mainloop()

if __name__ == "__main__":
    launch_gui()
